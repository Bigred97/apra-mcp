"""XLSX parsers for APRA quarterly publications.

APRA ships XLSX in two broad shapes:

1. **Long-format Database files** (General Insurance, Life Insurance — both
   current and historical). A single `Database` or `Data` sheet with one row
   per (date, data_item, dimensions...). Header row 1, ~14 columns. Ideal for
   filtering — the parser only needs to read it.

2. **Multi-tab presentation files** (ADI Centralised Publication, Super
   Fund-Level). Each Table N sheet is its own long-format slice, but headers
   land on row 3 with a title in row 1 and sometimes a units row in row 2.

3. **Transposed pivot tables** (Super Performance KeyStats, ADI Property
   Exposures). Rows are entity categories, column headers are time periods.
   Call `melt_transposed()` after `read_xlsx()` to normalise to long format.

We parse all three via `read_xlsx(sheet, header_row)` — the curated YAML pins
the exact row. The header is normalised so embedded newlines + extra
whitespace don't break exact column-name matching downstream.
"""
from __future__ import annotations

import calendar
import re as _re
import zipfile
from datetime import datetime
from io import BytesIO
from typing import Any, Callable

import openpyxl
import pandas as pd

_MONTH_ABBR: dict[str, int] = {
    "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
    "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12,
}
_MON_YYYY_RE = _re.compile(r"^([A-Za-z]{3})\s+(\d{4})$")


class ParseError(Exception):
    """Raised when an APRA resource can't be parsed."""


def read_xlsx(
    body: bytes,
    *,
    sheet: str,
    header_row: int,
    data_start_row: int | None = None,
    max_rows: int | None = None,
    period_source_column: str | None = None,
    start_period: str | None = None,
    end_period: str | None = None,
) -> pd.DataFrame:
    """Read one sheet from an XLSX as a DataFrame using streaming row iteration.

    Reads the sheet via openpyxl in read-only mode and accumulates rows in a
    list before handing them to pandas. This caps peak memory at the working
    set rather than the full sheet representation pandas would otherwise
    materialise — important for APRA's 7MB historical files (the General
    Insurance historical XLSX has ~142k rows, ~70MB peak under
    `pd.read_excel`, ~42MB peak under read-only iteration; row-skip filtering
    by period drops that to ~15MB).

    Args:
        body: raw bytes of the .xlsx file.
        sheet: sheet name (must exist).
        header_row: 1-indexed row containing column headers (matches Excel's
            row numbering and the convention used in curated YAMLs).
        data_start_row: 1-indexed first row of data. Defaults to header_row + 1.
            Set this when there are blank/spacer rows between header and data.
        max_rows: cap on data rows returned. Useful when APRA tables have
            trailing "Notes" rows after the data block.
        period_source_column: when set together with start_period or
            end_period, rows whose period-column value falls outside the
            inclusive range are skipped at iteration time. Avoids
            materialising rows we'd discard later in shaping. The column name
            is matched case-insensitively against the normalised header.
        start_period: optional inclusive lower bound. Accepts ISO date,
            year, year-month, or year-quarter forms (same conventions as
            `_expand_period_input` in shaping.py).
        end_period: optional inclusive upper bound, same accepted forms.

    Returns:
        DataFrame indexed 0..N-1, with column names normalised but otherwise
        identical to the source headers (renaming to plain-English aliases
        happens in shaping.py).
    """
    if not body:
        raise ParseError("empty XLSX body")
    if header_row < 1:
        raise ParseError(f"header_row must be 1-indexed (>=1), got {header_row}")
    if data_start_row is not None and data_start_row < header_row + 1:
        raise ParseError(
            f"data_start_row ({data_start_row}) must be > header_row ({header_row})"
        )

    period_predicate = _make_period_predicate(start_period, end_period)
    effective_start = data_start_row if data_start_row is not None else header_row + 1

    try:
        wb = openpyxl.load_workbook(BytesIO(body), read_only=True, data_only=True)
    except (KeyError, OSError, zipfile.BadZipFile, ValueError) as e:
        raise ParseError(f"could not parse XLSX (corrupt or truncated body): {e}") from e

    try:
        if sheet not in wb.sheetnames:
            raise ParseError(f"sheet {sheet!r} not found in workbook")
        ws = wb[sheet]

        header_cells: tuple[Any, ...] | None = None
        period_col_idx: int | None = None
        kept: list[tuple[Any, ...]] = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx == header_row:
                header_cells = row
                if period_source_column and period_predicate is not None:
                    target_lower = period_source_column.strip().lower()
                    for i, c in enumerate(row):
                        norm = _normalize_header(c) if isinstance(c, str) else c
                        if isinstance(norm, str) and norm.lower() == target_lower:
                            period_col_idx = i
                            break
                continue
            if row_idx < effective_start:
                continue
            if max_rows is not None and len(kept) >= max_rows:
                break
            if period_predicate is not None and period_col_idx is not None:
                if not period_predicate(row[period_col_idx]):
                    continue
            kept.append(row)
    finally:
        wb.close()

    if header_cells is None:
        raise ParseError(f"header row {header_row} not present in sheet {sheet!r}")

    # Normalise headers + give Unnamed:N names to blank cells, then disambiguate
    # collisions (".1", ".2", ...) so the resulting DataFrame matches
    # pd.read_excel's column-name behaviour. Several transposed APRA sheets
    # have a blank first header cell over the entity column.
    raw_cols = [_normalize_header(c) for c in header_cells]
    columns = _disambiguate_column_names(raw_cols)
    df = pd.DataFrame(kept, columns=columns)
    # Drop fully-blank trailing rows (where every cell is None / NaN). APRA
    # files routinely have a few empty rows after the data block.
    df = df.dropna(how="all").reset_index(drop=True)
    return df


def _disambiguate_column_names(cols: list[Any]) -> list[Any]:
    """Match pd.read_excel's column-name convention.

    - `None` / empty strings → `"Unnamed: <index>"` (0-based).
    - Duplicate names → suffix `.1`, `.2`, ... so each is unique.
    """
    blank_filled: list[Any] = []
    for i, c in enumerate(cols):
        if c is None:
            blank_filled.append(f"Unnamed: {i}")
        elif isinstance(c, str) and not c:
            blank_filled.append(f"Unnamed: {i}")
        else:
            blank_filled.append(c)
    seen: dict[Any, int] = {}
    out: list[Any] = []
    for c in blank_filled:
        if c in seen:
            seen[c] += 1
            out.append(f"{c}.{seen[c]}")
        else:
            seen[c] = 0
            out.append(c)
    return out


def _make_period_predicate(
    start_period: str | None, end_period: str | None
) -> Callable[[Any], bool] | None:
    """Return a fast row-skip predicate for a cell against an inclusive range.

    Bounds are expanded to ISO YYYY-MM-DD strings (start→01-01, end→12-31,
    quarter→QQ-end, etc.). Cell values may arrive as `datetime` (from typed
    date cells) or `str`; both are normalised to ISO strings before
    comparison. Returns None when no bounds are supplied.

    Comparison is lexicographic on the ISO string, which is correct for
    YYYY-MM-DD and well-defined for the partial forms APRA uses.
    """
    if not start_period and not end_period:
        return None
    start_iso = _expand_period_for_skip(start_period, bound="start") if start_period else None
    end_iso = _expand_period_for_skip(end_period, bound="end") if end_period else None

    def predicate(cell: Any) -> bool:
        if cell is None:
            return False
        if isinstance(cell, (pd.Timestamp, datetime)):
            try:
                iso = cell.strftime("%Y-%m-%d")
            except (ValueError, AttributeError):
                return False
        else:
            iso = str(cell).strip()
            if len(iso) >= 10 and iso[4:5] == "-" and iso[7:8] == "-":
                iso = iso[:10]
        if start_iso is not None and iso < start_iso:
            return False
        if end_iso is not None and iso > end_iso:
            return False
        return True

    return predicate


def _expand_period_for_skip(value: str, *, bound: str) -> str:
    """Expand a user-supplied period string to an ISO YYYY-MM-DD bound.

    Mirrors `shaping._expand_period_input` but kept local so parsing has no
    dependency on shaping. Accepts ISO dates, bare years, YYYY-MM,
    YYYY-Qx (case-insensitive). Anything else passes through.
    """
    if not value:
        return value
    s = value.strip()
    if len(s) == 10 and s[4] == "-" and s[7] == "-":
        return s
    if len(s) == 4 and s.isdigit():
        return f"{s}-01-01" if bound == "start" else f"{s}-12-31"
    if len(s) == 7 and s[4] == "-" and s[5] in ("Q", "q"):
        year = s[:4]
        try:
            q = int(s[6])
        except ValueError:
            return s
        if 1 <= q <= 4:
            if bound == "start":
                month = {1: "01", 2: "04", 3: "07", 4: "10"}[q]
                return f"{year}-{month}-01"
            end_md = {1: "03-31", 2: "06-30", 3: "09-30", 4: "12-31"}[q]
            return f"{year}-{end_md}"
    if len(s) == 7 and s[4] == "-" and s[5:].isdigit():
        try:
            m = int(s[5:7])
        except ValueError:
            return s
        if 1 <= m <= 12:
            if bound == "start":
                return f"{s}-01"
            last_day = calendar.monthrange(int(s[:4]), m)[1]
            return f"{s}-{last_day:02d}"
    return s


def _normalize_header(c):
    """Normalize an XLSX column header — collapse internal whitespace runs.

    APRA headers occasionally arrive with embedded newlines or duplicate
    spaces (`"Common Equity  Tier 1 capital"`, `"Total Tier 1\\ncapital"`).
    We collapse internal runs of whitespace (including newlines) to a single
    space and strip leading/trailing whitespace. Curated YAMLs spell columns
    in canonical (single-space) form.
    """
    if not isinstance(c, str):
        return c
    parts = c.replace("\r", " ").replace("\n", " ").split()
    return " ".join(parts)


def normalize_transposed_period(v: Any) -> str | None:
    """Convert a period label from a transposed APRA table to ISO YYYY-MM-DD.

    Handles:
    - pandas Timestamp / datetime → strftime('%Y-%m-%d')
    - 'Mar 2025' / 'Dec 2004' style month-year strings → last day of that month
    - Already-ISO 'YYYY-MM-DD' strings → pass through
    """
    if v is None:
        return None
    if isinstance(v, (pd.Timestamp, datetime)):
        try:
            return v.strftime("%Y-%m-%d")
        except (ValueError, AttributeError):
            return None
    s = str(v).strip()
    if not s:
        return None
    m = _MON_YYYY_RE.match(s)
    if m:
        month_name = m.group(1).capitalize()
        year = int(m.group(2))
        month_num = _MONTH_ABBR.get(month_name)
        if month_num:
            last_day = calendar.monthrange(year, month_num)[1]
            return f"{year}-{month_num:02d}-{last_day:02d}"
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return s[:10]
    return s


def _is_period_column_header(col: Any) -> bool:
    """Return True if a column name looks like a time-period label.

    Matches pandas Timestamps, datetimes, and 'Mon YYYY' strings.
    """
    if isinstance(col, (pd.Timestamp, datetime)):
        return True
    if not isinstance(col, str):
        return False
    return bool(_MON_YYYY_RE.match(col.strip()))


def melt_transposed(df: pd.DataFrame, entity_alias: str) -> pd.DataFrame:
    """Convert a pivot-table XLSX sheet to long-format DataFrame.

    APRA publishes some datasets in a transposed layout: rows are entity
    categories (fund types, property types) and column headers are time
    periods. This function melts them into (entity_alias, 'period', 'value')
    triples.

    The first column is treated as the entity dimension and renamed to
    `entity_alias`. All columns whose header passes `_is_period_column_header`
    are melted into period + value pairs. Other columns are discarded.

    Period headers are normalised to ISO YYYY-MM-DD via
    `normalize_transposed_period`.
    """
    if df.empty:
        return df

    entity_col = df.columns[0]
    period_cols = [c for c in df.columns[1:] if _is_period_column_header(c)]

    if not period_cols:
        return df

    df_work = df.copy()
    df_work = df_work.rename(columns={entity_col: entity_alias})
    df_work = df_work.dropna(subset=[entity_alias]).reset_index(drop=True)

    df_long = df_work.melt(
        id_vars=[entity_alias],
        value_vars=period_cols,
        var_name="period",
        value_name="value",
    )
    df_long["period"] = df_long["period"].apply(normalize_transposed_period)
    return df_long.reset_index(drop=True)


def drop_blank_rows(df: pd.DataFrame, key_columns: list[str]) -> pd.DataFrame:
    """Drop rows where every column in `key_columns` is NaN.

    Used to trim trailing footnote / blank rows that APRA sometimes leaves
    after the data block.
    """
    present = [c for c in key_columns if c in df.columns]
    if not present:
        return df
    keep_mask = ~df[present].isna().all(axis=1)
    return df.loc[keep_mask].reset_index(drop=True)
